import gzip
import pickle
import re
from typing import Union
from hashlib import md5

import openpyxl
from openpyxl_image_loader import SheetImageLoader

from .common import ASSET_DATA_DIR
from .effects import NUMBER_OF_EFFECTS, Effects
from .recipes import Recipe, PotionBases
from .recipes import Ingredients, Salts, NUMBER_OF_INGREDIENTS, NUMBER_OF_SALTS
from .recipes import Potion, IngredientNumList, SaltGrainList
from . import legendary as Legendary
from . import single_effect as SingleEffect

Numerical = Union[int, float]
Correction = {"Stentch": "Stench", "Rejuvination": "Rejuvenation"}


def to_int(x):
    if isinstance(x, Numerical):
        return int(x)
    if isinstance(x, str) and x.isnumeric():
        return int(x)
    return 0


def read_tome_recipes():

    tome = openpyxl.open(ASSET_DATA_DIR / "tome.xlsx", data_only=True)
    recipe_dump: set[Recipe] = set()

    ## Reading RecipeDump page
    tome_recipe_dump = tome["Recipe Dump"]
    recipe_dump_count = 0
    row = 2  # First meaningful row.
    while True:
        row += 1
        recipe_title = tome_recipe_dump.cell(row, 1).value
        if recipe_title is not None:
            recipe_title = str(recipe_title)
            legendary_pattern = re.compile(r"^([a-zA-Z]*) ([a-zA-Z]*)-(\d*)('?)$")
            single_effect_pattern = re.compile(r"^([a-zA-z]*) ?((?:[a-zA-z]*)?)('?)$")
            legendary_match = re.match(legendary_pattern, recipe_title)
            if legendary_match is not None:
                groups = legendary_match.groups()
                key = f"{groups[0]}{groups[1]}_{groups[2]}"
                if key in Correction:
                    key = Correction[key]
                potion = Legendary.__dict__[key]
                hidden = bool(groups[3])
            else:
                single_effect_match = re.match(single_effect_pattern, recipe_title)
                if single_effect_match is not None:
                    groups = single_effect_match.groups()
                    key = "".join([groups[0], groups[1]])
                    key = Correction.get(key, key)
                    tier = tome_recipe_dump.cell(row, 2).value
                    if isinstance(tier, Numerical):
                        tier = int(tier)
                        if not 1 <= tier <= 3:
                            print(f"Unexpected tier {tier} for single effects found at row {row}.")
                            continue
                    else:
                        print(f"Tier for single effect potion is not a number at row {row}.")
                        continue
                    key = ["Weak", "", "Strong"][tier - 1] + key
                    potion = SingleEffect.__dict__[key]
                    hidden = bool(groups[2])
                else:
                    # These recipes are temperarily removed because difficulty to makein real game.
                    print(f"Potion title matched neither legendary potions nor single effect potions at row {row}.")
                    continue
            base_title = tome_recipe_dump.cell(row, 3).value
            match base_title:
                case "Wa":
                    base = PotionBases.Water
                case "O":
                    base = PotionBases.Oil
                case "Wi":
                    base = PotionBases.Wine
                case _:
                    print(f"Unexpected baseTitle {base_title} assigned at row {row}.")
                    continue

            # to_int = lambda x: int(x) if x is not None else 0
            ingredient_num_list = IngredientNumList(to_int(tome_recipe_dump.cell(row, col).value) for col in range(20, 78))
            salt_grain_list = SaltGrainList(to_int(tome_recipe_dump.cell(row, col).value) for col in range(78, 83))
            plotter_link = tome_recipe_dump.cell(row, 84).value
            plotter_link = str(plotter_link) if plotter_link else ""
            discord_link = tome_recipe_dump.cell(row, 17).hyperlink
            discord_link = discord_link.target or "" if discord_link else ""
            hidden = hidden or not (bool(plotter_link) or bool(discord_link))
            if hidden:
                print(f"Info: row {row} is a hidden recipe.")
            recipe = Recipe(
                base,
                potion,
                ingredient_num_list,
                salt_grain_list,
                discord_link=discord_link,
                plotter_link=plotter_link,
                hidden=hidden,
            )
            if recipe in recipe_dump:
                print(f"row {row} records an identical recipe recorded before.")
            else:
                recipe_dump.add(recipe)
                recipe_dump_count += 1
        else:
            # last recipe row.
            break
    print(f"Reading {recipe_dump_count} recipes from recipe dump page.")

    ### Reading SaltySkirt page.
    tome_salty_skirt = tome["Salty Skirt"]
    col_letters = "ABCDE"
    image_loader = SheetImageLoader(tome_salty_skirt)
    salty_skirt_count = 0
    icon_md5_path = ASSET_DATA_DIR / "iconMD5s.pkl.gz"
    if not icon_md5_path.exists():
        effect_md5s = read_icon_md5()
        with gzip.open(icon_md5_path, "wb") as f:
            pickle.dump(effect_md5s, f)
    else:
        with gzip.open(icon_md5_path, "rb") as f:
            effect_md5s = pickle.load(f)

    for row in range(10, 204):  # currently not collecting backups.
        if image_loader.image_in(f"A{row}"):
            effect_tiers = [0] * NUMBER_OF_EFFECTS
            for col in range(5):
                if image_loader.image_in(f"{col_letters[col]}{row}"):
                    effect_icon = image_loader.get(f"{col_letters[col]}{row}")
                    effect_md5 = md5(pickle.dumps(effect_icon)).hexdigest()
                    effect = effect_md5s[effect_md5]
                    effect_tiers[effect] += 1
            potion = Potion(effect_tiers)
            # We can not retrieve the base from the page itself.
            _ingredient_nums = [0] * NUMBER_OF_INGREDIENTS
            _ingredient_nums[Ingredients.PhantomSkirt] += to_int(tome_salty_skirt.cell(row, 16).value)
            _ingredient_nums[Ingredients.GraveTruffle] += to_int(tome_salty_skirt.cell(row, 17).value)
            _ingredient_nums[Ingredients.Watercap] += to_int(tome_salty_skirt.cell(row, 18).value)
            _ingredient_nums[Ingredients.Goldthorn] += to_int(tome_salty_skirt.cell(row, 19).value)
            _ingredient_nums[Ingredients.Boombloom] += to_int(tome_salty_skirt.cell(row, 20).value)
            _ingredient_nums[Ingredients.RainbowCap] += to_int(tome_salty_skirt.cell(row, 21).value)
            _ingredient_nums[Ingredients.FrostSapphire] += to_int(tome_salty_skirt.cell(row, 22).value)
            ingredient_num_list = IngredientNumList(_ingredient_nums)
            _salt_grains = [0] * NUMBER_OF_SALTS
            _salt_grains[Salts.Moon] += to_int(tome_salty_skirt.cell(row, 13).value)
            _salt_grains[Salts.Sun] += to_int(tome_salty_skirt.cell(row, 14).value)
            _salt_grains[Salts.Life] += to_int(tome_salty_skirt.cell(row, 15).value)
            salt_grain_list = SaltGrainList(_salt_grains)
            plotter_link = tome_salty_skirt.cell(row, 7).hyperlink
            plotter_link = plotter_link.target or "" if plotter_link else ""
            recipe = Recipe(PotionBases.Unknown, potion, ingredient_num_list, salt_grain_list, plotter_link=plotter_link)
            if recipe not in recipe_dump:
                recipe_dump.add(recipe)
                salty_skirt_count += 1
        else:
            continue
    print(f"Completed reading {salty_skirt_count} additional recipes from salty skirt page.")
    return recipe_dump


class CustomerRequest:
    def __init__(self, idx: int, name: str, requested_effects: list[Effects], text: str, carma: int, story_line=""):
        self.idx = idx
        self.name = name
        self.text = text
        self.requested_effects = requested_effects
        self.carma = carma
        self.story_line = story_line

    def __str__(self):
        return f"idx: {self.idx}\nname: {self.name}\nrequested_effects: {self.requested_effects}\ntext: {self.text}\ncarma: {self.carma}\n"

    def __repr__(self) -> str:
        return self.__str__()


def read_tome_customers_requests():
    tome = openpyxl.open(ASSET_DATA_DIR / "tome.xlsx", data_only=True)
    customers_requests = tome["Customer Requests"]
    _customers_requests = []
    _story_lines = []
    _read_count = 0
    _row_idx = 1  # Header row
    _read_requests = re.compile(r" *[( ]?(\w+)[ )]?")
    _read_story_line = re.compile(r"^(\w*)_\d_\w*$")
    while True:
        _row_idx += 1
        _name_text = customers_requests.cell(_row_idx, 1).value
        if _name_text is not None:
            _name = str(_name_text)
            _story_line_match = re.match(_read_story_line, _name)
            if _story_line_match is not None:
                _story_line = _story_line_match.group(1)
                if _story_line not in _story_lines:
                    _story_lines.append(_story_line)
            else:
                _story_line = ""
            _request_text = str(customers_requests.cell(_row_idx, 2).value)
            _requested_effects_text = str(customers_requests.cell(_row_idx, 3).value)
            _requested_effects = re.findall(_read_requests, _requested_effects_text)
            if _requested_effects is not None:
                _requested_effects = [Effects[effect] for effect in _requested_effects]
                _read_count += 1
            else:
                print(f"Parsing error at row {_row_idx}.")
                continue
            _carma_text = customers_requests.cell(_row_idx, 4).value
            _carma = int(float(str(_carma_text))) if _carma_text else 0
            _customers_requests.append(CustomerRequest(_read_count, _name, _requested_effects, _request_text, _carma, _story_line))
        else:
            break
    print(f"Completed reading {_read_count} customers requests from customers requests page.")

    return _customers_requests, _story_lines


def read_tome_effect_compatibilties():
    tome = openpyxl.open(ASSET_DATA_DIR / "tome.xlsx", data_only=True)
    _page_name = r"Compatible Effects (Groups)"
    page = tome[_page_name]
    image_loader = SheetImageLoader(page)
    icon_md5_path = ASSET_DATA_DIR / "iconMD5s.pkl.gz"
    # load effect icon hash value.
    if not icon_md5_path.exists():
        effect_md5s = read_icon_md5()
        with gzip.open(icon_md5_path, "wb") as f:
            pickle.dump(effect_md5s, f)
    else:
        with gzip.open(icon_md5_path, "rb") as f:
            effect_md5s = pickle.load(f)

    effect_list: list[Effects] = []
    for row in range(4, 45):
        assert image_loader.image_in(f"C{row}")
        effect_image = image_loader.get(f"C{row}")
        # print(row)
        effect_md5 = md5(pickle.dumps(effect_image)).hexdigest()
        effect = effect_md5s[effect_md5]
        effect_list.append(effect)
    print(effect_list)
    _compatibilitie_matrix: list[list[int]] = [[0 for _ in range(41)] for _ in range(41)]
    for i in range(41):
        main_effect = effect_list[i]
        for j in range(41):
            secondary_effect = effect_list[j]
            value = page.cell(i + 4, j + 4).value
            # if isinstance(value, float):
            #     _compatibilitie_matrix[i][j] = int(value)
            # else:
            #     _compatibilitie_matrix[i][j] = -1
            _compatibilitie_matrix[main_effect][secondary_effect] = int(value) if isinstance(value, float) else -1
    with gzip.open(ASSET_DATA_DIR / "Compatibility.pkl.gz", "wb") as f:
        pickle.dump(_compatibilitie_matrix, f)
    return _compatibilitie_matrix


def read_icon_md5():
    from .common import EXAMPLE_EFFECT_ICON_ROWS_SALTY_SKIRT, EXAMPLE_EFFECT_ICON_ROWS_COMPATIBILITY

    tome = openpyxl.open(ASSET_DATA_DIR / "tome.xlsx", data_only=True)
    tome_salty_skirt = tome["Salty Skirt"]
    tome_compatible_effects = tome["Compatible Effects (Groups)"]
    image_loader_salty_skirt = SheetImageLoader(tome_salty_skirt)
    image_loader_compatible_effects = SheetImageLoader(tome_compatible_effects)

    ### load icon from Salty Skirt page.
    example_icon_rows = EXAMPLE_EFFECT_ICON_ROWS_SALTY_SKIRT
    icon_md5 = {}
    for index, row in enumerate(example_icon_rows):
        image = image_loader_salty_skirt.get(f"A{row}")
        icon_md5[md5(pickle.dumps(image)).hexdigest()] = index

    ### load icon from Compatible Effects (Groups) page.
    example_icon_rows = EXAMPLE_EFFECT_ICON_ROWS_COMPATIBILITY
    for index, row in enumerate(example_icon_rows):
        image = image_loader_compatible_effects.get(f"C{row}")
        icon_md5[md5(pickle.dumps(image)).hexdigest()] = index

    return icon_md5


def update_icon_md5():
    icon_md5 = read_icon_md5()
    with gzip.open(ASSET_DATA_DIR / "iconMD5s.pkl.gz", "wb") as f:
        pickle.dump(icon_md5, f)


if __name__ == "__main__":
    # print(read_icon_md5())
    # read_tome_recipes()
    # _data = read_tome_customers_requests()
    #
    md5s = pickle.load(gzip.open(ASSET_DATA_DIR / "iconMD5s.pkl.gz", "rb"))
    print(md5s)

    result = read_tome_effect_compatibilties()
    print(result)
