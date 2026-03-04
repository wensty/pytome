import re

import openpyxl

from .common import ASSET_DATA_DIR
from .effects import Effects


class CustomerRequest:
    def __init__(self, idx: int, name: str, requested_effects: list[Effects], text: str, carma: int, story_line=""):
        self.idx = idx
        self.name = name
        self.text = text
        self.requested_effects = requested_effects
        self.carma = carma
        self.story_line = story_line

    def __str__(self):
        return f"idx: {self.idx}\nname: {self.name}\nrequested_effects: {self.requested_effects}\ntext: {self.text}\ncarma: {self.carma}\n"

    def __repr__(self) -> str:
        return self.__str__()


def read_tome_customers_requests():
    tome = openpyxl.open(ASSET_DATA_DIR / "tome.xlsx", data_only=True)
    customers_requests = tome["Customer Requests"]
    _customers_requests = []
    _story_lines = []
    _read_count = 0
    _row_idx = 1  # Header row
    _read_requests = re.compile(r" *[( ]?(\w+)[ )]?")
    _read_story_line = re.compile(r"^(\w*)_\d_\w*$")
    while True:
        _row_idx += 1
        _name_text = customers_requests.cell(_row_idx, 1).value
        if _name_text is not None:
            _name = str(_name_text)
            _story_line_match = re.match(_read_story_line, _name)
            if _story_line_match is not None:
                _story_line = _story_line_match.group(1)
                if _story_line not in _story_lines:
                    _story_lines.append(_story_line)
            else:
                _story_line = ""
            _request_text = str(customers_requests.cell(_row_idx, 2).value)
            _requested_effects_text = str(customers_requests.cell(_row_idx, 3).value)
            _requested_effects = re.findall(_read_requests, _requested_effects_text)
            if _requested_effects is not None:
                _requested_effects = [Effects[effect] for effect in _requested_effects]
                _read_count += 1
            else:
                print(f"Parsing error at row {_row_idx}.")
                continue
            _carma_text = customers_requests.cell(_row_idx, 4).value
            _carma = int(float(str(_carma_text))) if _carma_text else 0
            _customers_requests.append(CustomerRequest(_read_count, _name, _requested_effects, _request_text, _carma, _story_line))
        else:
            break
    print(f"Completed reading {_read_count} customers requests from customers requests page.")

    return _customers_requests, _story_lines


if __name__ == "__main__":
    _data = read_tome_customers_requests()
    print(_data)
