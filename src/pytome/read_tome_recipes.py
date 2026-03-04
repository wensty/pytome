import pickle
import re
from hashlib import md5
from typing import Union

import openpyxl
from openpyxl_image_loader import SheetImageLoader

from .common import ASSET_DATA_DIR, effect_md5s
from .effects import NUMBER_OF_EFFECTS
from .recipes import Recipe, PotionBases, Comment, CommentType, LinkType, RecipeLink
from .recipes import Ingredients, Salts, NUMBER_OF_INGREDIENTS, NUMBER_OF_SALTS
from .recipes import Potion, IngredientNumList, SaltGrainList
from . import legendary as Legendary
from . import single_effect as SingleEffect

Numerical = Union[int, float]
Correction = {"Stentch": "Stench", "Rejuvination": "Rejuvenation"}


def to_int(x):
    if isinstance(x, Numerical):
        return int(x)
    if isinstance(x, str) and x.isnumeric():
        return int(x)
    return 0


def read_tome_recipes():
    tome = openpyxl.open(ASSET_DATA_DIR / "tome.xlsx", data_only=True)
    recipe_dump: set[Recipe] = set()
    comments: list[Comment] = []
    links: list[RecipeLink] = []

    ## Reading RecipeDump page
    tome_recipe_dump = tome["Recipe Dump"]
    recipe_dump_count = 0
    row = 2  # First meaningful row.
    while True:
        row += 1
        recipe_title = tome_recipe_dump.cell(row, 1).value
        if recipe_title is not None:
            recipe_title = str(recipe_title)
            legendary_pattern = re.compile(r"^([a-zA-Z]*) ([a-zA-Z]*)-(\d*)('?)$")
            single_effect_pattern = re.compile(r"^([a-zA-z]*) ?((?:[a-zA-z]*)?)('?)$")
            legendary_match = re.match(legendary_pattern, recipe_title)
            if legendary_match is not None:
                groups = legendary_match.groups()
                key = f"{groups[0]}{groups[1]}_{groups[2]}"
                if key in Correction:
                    key = Correction[key]
                potion = Legendary.__dict__[key]
                # Only the explicit special marker denotes hidden recipes.
                hidden = bool(groups[3])
            else:
                single_effect_match = re.match(single_effect_pattern, recipe_title)
                if single_effect_match is not None:
                    groups = single_effect_match.groups()
                    key = "".join([groups[0], groups[1]])
                    key = Correction.get(key, key)
                    tier = tome_recipe_dump.cell(row, 2).value
                    if isinstance(tier, Numerical):
                        tier = int(tier)
                        if not 1 <= tier <= 3:
                            print(f"Unexpected tier {tier} for single effects found at row {row}.")
                            continue
                    else:
                        print(f"Tier for single effect potion is not a number at row {row}.")
                        continue
                    key = ["Weak", "", "Strong"][tier - 1] + key
                    potion = SingleEffect.__dict__[key]
                    # Only the explicit special marker denotes hidden recipes.
                    hidden = bool(groups[2])
                else:
                    # These recipes are temperarily removed because difficulty to makein real game.
                    print(f"Potion title matched neither legendary potions nor single effect potions at row {row}.")
                    continue
            base_title = tome_recipe_dump.cell(row, 3).value
            match base_title:
                case "Wa":
                    base = PotionBases.Water
                case "O":
                    base = PotionBases.Oil
                case "Wi":
                    base = PotionBases.Wine
                case _:
                    print(f"Unexpected baseTitle {base_title} assigned at row {row}.")
                    continue

            ingredient_num_list = IngredientNumList(to_int(tome_recipe_dump.cell(row, col).value) for col in range(20, 78))
            salt_grain_list = SaltGrainList(to_int(tome_recipe_dump.cell(row, col).value) for col in range(78, 83))
            plotter_link = tome_recipe_dump.cell(row, 84).value
            plotter_link = str(plotter_link) if plotter_link else ""
            discord_link = tome_recipe_dump.cell(row, 17).hyperlink
            discord_link = discord_link.target or "" if discord_link else ""
            if hidden:
                print(f"Info: row {row} is a hidden recipe.")
            recipe = Recipe(
                base,
                potion,
                ingredient_num_list,
                salt_grain_list,
                hidden=hidden,
            )
            if recipe in recipe_dump:
                print(f"row {row} records an identical recipe recorded before.")
            else:
                recipe_dump.add(recipe)
                recipe_dump_count += 1
            if plotter_link:
                links.append(RecipeLink(recipe, LinkType.Plotter, plotter_link))
            if discord_link:
                links.append(RecipeLink(recipe, LinkType.Discord, discord_link))
            # read comments. This page contains only plotter comments.
            comment = tome_recipe_dump.cell(row, 16).comment
            if comment is not None:
                comment_text = comment.text
                comment_author = comment.author if comment.author != "None" else "Anonymous"
                comment = Comment(recipe, CommentType.Plotter, comment_author, comment_text)
                comments.append(comment)
        else:
            # last recipe row.
            break
    print(f"Reading {recipe_dump_count} recipes from recipe dump page.")

    ### Reading SaltySkirt page.
    tome_salty_skirt = tome["Salty Skirt"]
    col_letters = "ABCDE"
    image_loader = SheetImageLoader(tome_salty_skirt)
    salty_skirt_count = 0

    for row in range(10, 204):  # currently not collecting backups.
        if image_loader.image_in(f"A{row}"):  # recipe row.
            effect_tiers = [0] * NUMBER_OF_EFFECTS
            for col in range(5):
                if image_loader.image_in(f"{col_letters[col]}{row}"):
                    effect_icon = image_loader.get(f"{col_letters[col]}{row}")
                    effect_md5 = md5(pickle.dumps(effect_icon)).hexdigest()
                    effect = effect_md5s[effect_md5]
                    effect_tiers[effect] += 1
            potion = Potion(effect_tiers)
            # We can not retrieve the base from the page itself.
            _ingredient_nums = [0] * NUMBER_OF_INGREDIENTS
            _ingredient_nums[Ingredients.PhantomSkirt] += to_int(tome_salty_skirt.cell(row, 16).value)
            _ingredient_nums[Ingredients.GraveTruffle] += to_int(tome_salty_skirt.cell(row, 17).value)
            _ingredient_nums[Ingredients.Watercap] += to_int(tome_salty_skirt.cell(row, 18).value)
            _ingredient_nums[Ingredients.Goldthorn] += to_int(tome_salty_skirt.cell(row, 19).value)
            _ingredient_nums[Ingredients.Boombloom] += to_int(tome_salty_skirt.cell(row, 20).value)
            _ingredient_nums[Ingredients.RainbowCap] += to_int(tome_salty_skirt.cell(row, 21).value)
            _ingredient_nums[Ingredients.FrostSapphire] += to_int(tome_salty_skirt.cell(row, 22).value)
            ingredient_num_list = IngredientNumList(_ingredient_nums)
            _salt_grains = [0] * NUMBER_OF_SALTS
            _salt_grains[Salts.Moon] += to_int(tome_salty_skirt.cell(row, 13).value)
            _salt_grains[Salts.Sun] += to_int(tome_salty_skirt.cell(row, 14).value)
            _salt_grains[Salts.Life] += to_int(tome_salty_skirt.cell(row, 15).value)
            salt_grain_list = SaltGrainList(_salt_grains)
            plotter_link = tome_salty_skirt.cell(row, 7).hyperlink
            plotter_link = plotter_link.target or "" if plotter_link else ""
            recipe = Recipe(PotionBases.Unknown, potion, ingredient_num_list, salt_grain_list, hidden=False)
            if recipe not in recipe_dump:
                recipe_dump.add(recipe)
                salty_skirt_count += 1
            if plotter_link:
                links.append(RecipeLink(recipe, LinkType.Plotter, plotter_link))
            # read comments.
            comment_plotter = tome_salty_skirt.cell(row, 7).comment
            if comment_plotter is not None:
                comment_text = comment_plotter.text
                comment_author = comment_plotter.author if comment_plotter.author else "Anonymous"
                comment = Comment(recipe, CommentType.Plotter, comment_author, comment_text)
                comments.append(comment)

            recipe_note = tome_salty_skirt.cell(row, 8).value
            if recipe_note:
                comment = Comment(recipe, CommentType.Note, "Anonymous", str(recipe_note))
                comments.append(comment)

            recipe_note_comment = tome_salty_skirt.cell(row, 8).comment
            if recipe_note_comment is not None:
                comment_text = recipe_note_comment.text
                comment_author = recipe_note_comment.author if recipe_note_comment.author != "None" else "Anonymous"
                comment = Comment(recipe, CommentType.NoteComment, comment_author, comment_text)
                comments.append(comment)

            recipe_note_link = tome_salty_skirt.cell(row, 8).hyperlink
            if recipe_note_link:
                comment = Comment(recipe, CommentType.NoteLink, "Anonymous", recipe_note_link.target or "")
                comments.append(comment)

            recipe_moon_salt_comment = tome_salty_skirt.cell(row, 13).comment
            if recipe_moon_salt_comment is not None:
                comment_text = recipe_moon_salt_comment.text
                comment_author = recipe_moon_salt_comment.author if recipe_moon_salt_comment.author != "None" else "Anonymous"
                comment = Comment(recipe, CommentType.MoonSalt, comment_author, comment_text)
                comments.append(comment)

            recipe_sun_salt_comment = tome_salty_skirt.cell(row, 14).comment
            if recipe_sun_salt_comment is not None:
                comment_text = recipe_sun_salt_comment.text
                comment_author = recipe_sun_salt_comment.author if recipe_sun_salt_comment.author != "None" else "Anonymous"
                comment = Comment(recipe, CommentType.SunSalt, comment_author, comment_text)
                comments.append(comment)

            recipe_other_comment = tome_salty_skirt.cell(row, 22).comment
            if recipe_other_comment is not None:
                comment_text = recipe_other_comment.text
                comment_author = recipe_other_comment.author if recipe_other_comment.author != "None" else "Anonymous"
                comment = Comment(recipe, CommentType.Other, comment_author, comment_text)
                comments.append(comment)
        else:
            continue
    print(f"Completed reading {salty_skirt_count} additional recipes from salty skirt page.")
    print(f"Read {len(comments)} comments in total.")
    print(f"Read {len(links)} links in total.")
    return recipe_dump, comments, links


if __name__ == "__main__":
    recipe_dump, comment_dump, link_dump = read_tome_recipes()
    print(f"Read {len(recipe_dump)} recipes, {len(comment_dump)} comments and {len(link_dump)} links.")
    # print(comment_dump)
    # with open("recipe_dump.pkl", "wb") as f:
    #     pickle.dump(recipe_dump, f)
    # with open("comment_dump.pkl", "wb") as f:
    #     pickle.dump(comment_dump, f)
